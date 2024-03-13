from openpyxl import load_workbook


def read_profiles(filename="profiles.xlsx") -> dict:
	profiles = {}
	wb = load_workbook(filename=filename)
	sheet = wb.active
	for n, col in enumerate(sheet[1]):
		if col.value == "id":
			id_col = n
		if col.value == "name":
			name_col = n
	for row in sheet.iter_rows(min_row=2, values_only=True):
		id = row[id_col]
		name = row[name_col]
		profiles[id] = name
	return profiles